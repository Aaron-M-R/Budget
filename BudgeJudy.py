import math
import io
import os
import json
import calendar
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook

import csv
from csv import DictReader




#### Functions ####

# Returns date in MM-DD-YYYY format
def dater(date):
    day = datetime.strptime(str(date)[:10], "%Y-%m-%d")
    return day

# Keep charges within entered date range
def date_range(date):
    start = pd.to_datetime(first_fiscal, format="%Y-%m-%d")
    end = pd.to_datetime(last_fiscal, format="%Y-%m-%d")
    this = pd.to_datetime(date, format="%Y-%m-%d")
    return (end-this).days>=0 and (this-start).days>=0

# Takes in type and description of charge and returns a tri tuple of info
def description_and_amount(Type, Description, Amount):
    for category in list(categories.items()):
        for subcategory in list(category[1].items()):
            for code in subcategory[1]:
                if code.lower() in Description.lower() or code.lower() in Type.lower():

                    # return category, subcategory and charge description
                    return [category[0], subcategory[0], code]

    return ["Miscellaneous", "Miscellaneous", code]

# Combines withdrawal and deposit columns into one 'amount' column
def amount(df):
    if not math.isnan(df[5]):
        return -df[5]
    if not math.isnan(df[6]):
        return df[6]
    else:
        return 0

# Create cumulative lineplot of spending by category over time
def cat_plot(cats, df):
    new_df = df.copy()
    
    if isinstance(cats, str):
        cats = [cats]
    if isinstance(cats, list):
        new_df = new_df[new_df['Category'].isin(cats)]
        new_df['Amount'] = new_df['Amount'].apply(lambda x: -x)

        # Create a pivot table for cumulative sum of amount by category and date
        pivot_table = new_df.pivot_table(index='Date', columns='Category', values='Amount', aggfunc='sum', sort=False).fillna(0)
        cumulative_df = pivot_table.cumsum()

        # Plotting
        cumulative_df.plot(kind='line', linestyle='-', figsize=(10, 6))
        plt.title('Cumulative Amount Spent on Each Category Over Time')
        plt.xlabel('Date')
        plt.ylabel('Cumulative Spending')
        plt.legend(title='Category')
        plt.grid(True)
        plt.show()
    else:
        print("Input must be a list.")




#### Loading Stored Data ####

# User enters name of excel workbook
print("To begin, enter the name of your excel workbook.")
wb_name = input()
if "." not in wb_name:
    wb_name = f"{wb_name}.xlsx"

# Read in spending spreadsheet
sheetpath = Path('Data') / wb_name
df = pd.read_excel(sheetpath)

# Read in category descriptions of charges
filepath = Path('Data') / 'category_descriptions.json'
categories = json.loads(filepath.read_text())




#### User Enters Start and End Dates ####

startday, startmonth, startyear = 0, 0, 0
endyear, endmonth, endday = 0, 0, 0

# Start date info
print("Enter the starting date for your budget:")

# Collect starting year
while startyear == 0:
    print("Starting Year (YYYY)")
    year = input().strip()
    if len(year) == 2:
        startyear = "20" + year
    elif len(year) == 4:
        startyear = year
    else:
        print("Sorry! Your date was not in (YYYY) format. Please try again.")

# Collect starting month
while startmonth==0:
    print("Starting Month (MM)")
    month = input().strip()
    if len(month) == 1:
        month = "0" + month
    if 1<=int(month)<=12:
        startmonth = month

    else:
        print("Sorry! "+month+" isn\"t a valid month. Please try again.")

# Collect starting day
while startday == 0:
    print("Starting Day (DD)")
    day = input().strip()
    if len(day) == 1:
        day = "0" + day
    if 1 <= int(day) <= 31:
        startday = day
    else:
        print("Sorry! Your day is not in the range of the days in the month. Please try again.")

# End date info
print("Enter the ending date for your budget:")
print("Do you want to see your spending up to the present? (Y/N)")
answer = input()
if "y" in answer.lower():
    last_fiscal = str(date.today())
else:

    # Collect ending year
    while endyear == 0:
        print("Final Year (YYYY)")
        year = input().strip()
        if len(year) == 2:
            endyear = "20" + year
        elif len(year) == 4:
            endyear = year
        else:
            print("Sorry! Your date was not in (YYYY) format. Please try again.")

    # Collect ending month
    while endmonth==0:
        print("Final Month (MM)")
        month = input().strip()
        if len(month) == 1:
            month = "0" + month
        if 1<=int(month)<=12:
            endmonth = month
        else:
            print("Sorry! "+month+" isn\"t a valid month. Please try again.")

    # Collect ending day
    while endday == 0:
        print("Final Day (DD)")
        day = input().strip()
        if len(day) == 1:
            day = "0" + day
        if 1 <= int(day) <= 31:
            endday = day
        else:
            print("Sorry! Your day is not in the range of the days in the month. Please try again.")

    last_fiscal = "{}-{}-{}".format(endyear, endmonth, endday)
first_fiscal = "{}-{}-{}".format(startyear, startmonth, startday)

# The range of dates selected by the user
budget_days = pd.date_range(start=first_fiscal, end=last_fiscal)




#### Spreadsheet transformations ####

# Crop the rows according to desired date range
df['Date'] = df['Date'].apply(dater)

# Combine Deposit and Withdrawal columns into one Amount column
df['Amount'] = df.apply(amount, axis=1)

# Select date range given earlier
df = df[df['Date'].apply(date_range)]

# Add columns that describe charge (Category, Subcategory, Code)
df = df.assign(description = df.apply(lambda x: description_and_amount(x["Type"], x["Description"], x["Amount"]), axis=1))
charge_info = pd.DataFrame(df.description.tolist(), index=df.index, columns=['Category', 'Subcategory', 'Code'])
df = pd.concat([df, charge_info], axis=1)

# Rename columns
df = df[['Date', 'Category', 'Subcategory', 'Code', 'Amount', 'RunningBalance']]
df = df.reset_index().iloc[:, 1:]

# Create Year, Month and Day columns
df = df.assign(Year=df['Date'].apply(lambda x: int(str(x)[:4])),
               Month=df['Date'].apply(lambda x: int(str(x)[5:7])),
               Day=df['Date'].apply(lambda x: int(str(x)[8:10])))

# Create spending table grouped by month
totals = df.drop(columns=['Date', 'Code', 'RunningBalance', 'Day'])
totals = totals.groupby(['Year', 'Month', 'Category', 'Subcategory']).sum()
totals = totals.reset_index(level=1, drop=False)
totals = totals.assign(Month=totals['Month'].apply(lambda x: calendar.month_name[x]))
totals = totals.pivot_table(index=['Category', 'Subcategory'], columns='Month', values='Amount', aggfunc='sum', sort=False).fillna(0)
totals = totals.assign(Total=totals.sum(axis=1))

# Creat a dictionary for each category
list_of_dicts = [{item[0].capitalize(): list(item[1].keys())} for item in categories.items()]

# Extract keys and values from the list of dictionaries
keys = [list(d.keys())[0] for d in list_of_dicts]
values = [list(d.values())[0] for d in list_of_dicts]

# Create a MultiIndex from the keys and values
multi_index = pd.MultiIndex.from_tuples([(key, val) for key, val_list in zip(keys, values) for val in val_list], names=['Category', 'Subcategory'])

# Create an empty DataFrame and fill it with values from totals
empty_df = pd.DataFrame(index=multi_index, columns=[])
totals = empty_df.combine_first(totals).fillna(0)




#### Create Spreadsheet ####

# Load the existing Excel workbook
workbook = openpyxl.load_workbook(sheetpath)

# Get the old sheet
oldsheet_name = workbook.sheetnames[0]
oldsheet = workbook[oldsheet_name]

# Changes name of original sheet #
print(f"Would you like to rename this sheet from '{oldsheet_name}'? (Y/N))")
answer = input()
if 'y' in answer.lower():
    print("Provide the new name or press enter to keep original name")
    sheet_newname = input()
    if len(rename)>0:
        oldsheet.title = sheet_newname

# Save the changes to the workbook
workbook.save(sheetpath)

# Name new excel sheet according to user input
print("What would you like to call your new sheet?")
newsheet_name = input()

# Open the existing Excel file
with pd.ExcelWriter(sheetpath, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:

    # Write the DataFrame to a new sheet
    totals.to_excel(writer, sheet_name=newsheet_name, index=True)




#### Plotting ####

# Category Bar Plot #
plot_df = totals.copy()[['Total']].reset_index(level=1, drop=True)
plot_df = plot_df.reset_index().groupby('Category').sum()
plot_df['Total'] = plot_df['Total'].apply(lambda x: -x)
plot_df.plot(kind='bar', legend=None, figsize=(10, 6))
plt.title('Total Amount Spent on Each Category')
plt.xlabel('Category')
plt.ylabel('Total Amount')
plt.xticks(rotation=45, ha='right')
plt.show()

# Subcategory Bar Plot #
subplot_df = totals.copy()[['Total']].reset_index(level=0, drop=True)
subplot_df['Total'] = subplot_df['Total'].apply(lambda x: -x)
subplot_df.plot(kind='bar', legend=None, figsize=(10, 6))
plt.title('Total Amount Spent on Each Subcategory')
plt.xlabel('Category')
plt.ylabel('Total Amount')
plt.xticks(rotation=45, ha='right')
plt.show()

# Line Plot #
print("Would you like to visualize your spending over time? (Y/N)")
answer = input()
if 'y' in answer.lower():
    plotting = True
    while plotting:
        print("Enter the categories you want to see. Answer \'all\' to see every category")
        vistype = input()
        if vistype == 'all':
            viscats = list(df['Category'].unique())
        else:
            viscats = [word.strip() for word in vistype.strip('[]').split(',')]
        cat_plot(viscats, df)
        print("Would you like to plot again? (Y/N)")
        answer = input()
        if 'y' not in answer.lower():
            plotting = False


